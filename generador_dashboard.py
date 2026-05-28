from datetime import datetime

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MAX_TOP_CATEGORIAS = 12
MAX_FILAS_EXPORT_BASE = 50000


def leer_archivo(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        return pd.read_csv(archivo)
    return pd.read_excel(archivo)


def es_columna_vacia(serie):
    return serie.isna().all() or serie.astype(str).str.strip().isin(["", "nan", "None"]).all()


def detectar_columnas_fecha(df):
    columnas_fecha = []

    for columna in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[columna]):
            columnas_fecha.append(columna)
            continue

        muestra = df[columna].dropna().astype(str).head(300)
        if muestra.empty:
            continue

        fechas = parsear_fechas(muestra)
        ratio = fechas.notna().mean()

        if ratio >= 0.65:
            columnas_fecha.append(columna)

    return columnas_fecha


def parsear_fechas(serie):
    try:
        return pd.to_datetime(serie, errors="coerce", dayfirst=True, format="mixed")
    except TypeError:
        return pd.to_datetime(serie, errors="coerce", dayfirst=True)


def detectar_columnas_numericas(df):
    numericas = []

    for columna in df.columns:
        if pd.api.types.is_numeric_dtype(df[columna]):
            numericas.append(columna)
            continue

        muestra = df[columna].dropna().astype(str).head(300)
        if muestra.empty:
            continue

        limpia = (
            muestra.str.replace("€", "", regex=False)
            .str.replace("$", "", regex=False)
            .str.replace("£", "", regex=False)
            .str.replace(" ", "", regex=False)
        )
        limpia = limpia.apply(normalizar_numero_texto)
        convertida = pd.to_numeric(limpia, errors="coerce")

        if convertida.notna().mean() >= 0.75:
            numericas.append(columna)

    return numericas


def normalizar_numero_texto(valor):
    texto = str(valor)
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(",", ".")
    return texto


def serie_numerica(serie):
    if pd.api.types.is_numeric_dtype(serie):
        return pd.to_numeric(serie, errors="coerce")

    limpia = (
        serie.astype(str)
        .str.replace("€", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("£", "", regex=False)
        .str.replace(" ", "", regex=False)
    )
    limpia = limpia.apply(normalizar_numero_texto)
    return pd.to_numeric(limpia, errors="coerce")


def columnas_categoricas(df, columnas_fecha, columnas_numericas):
    excluidas = set(columnas_fecha) | set(columnas_numericas)
    categorias = []

    for columna in df.columns:
        if columna in excluidas:
            continue
        if es_columna_vacia(df[columna]):
            continue

        nunique = df[columna].nunique(dropna=True)
        if 1 < nunique <= max(80, len(df) * 0.35):
            categorias.append(columna)

    return categorias


def elegir_columna_importe(df, columnas_numericas):
    palabras_clave = ["importe", "facturacion", "facturación", "ventas", "revenue", "total", "amount", "precio"]

    for columna in columnas_numericas:
        nombre = str(columna).lower()
        if any(palabra in nombre for palabra in palabras_clave):
            return columna

    if columnas_numericas:
        return columnas_numericas[0]

    return None


def construir_calidad(df):
    filas = []
    total = len(df)

    for columna in df.columns:
        vacios = int(df[columna].isna().sum())
        unicos = int(df[columna].nunique(dropna=True))
        filas.append(
            {
                "Columna": columna,
                "Tipo_Detectado": tipo_legible(df[columna]),
                "Valores_Vacios": vacios,
                "Porcentaje_Vacios": round((vacios / total) * 100, 2) if total else 0,
                "Valores_Unicos": unicos,
            }
        )

    return pd.DataFrame(filas)


def tipo_legible(serie):
    if pd.api.types.is_datetime64_any_dtype(serie):
        return "Fecha"
    if pd.api.types.is_integer_dtype(serie):
        return "Número entero"
    if pd.api.types.is_float_dtype(serie):
        return "Número decimal"
    if pd.api.types.is_bool_dtype(serie):
        return "Sí/No"

    muestra = serie.dropna().astype(str).head(300)
    if muestra.empty:
        return "Vacío"

    fechas = parsear_fechas(muestra)
    if fechas.notna().mean() >= 0.65:
        return "Fecha"

    limpia = (
        muestra.str.replace("€", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("£", "", regex=False)
        .str.replace(" ", "", regex=False)
    )
    limpia = limpia.apply(normalizar_numero_texto)
    numeros = pd.to_numeric(limpia, errors="coerce")
    if numeros.notna().mean() >= 0.75:
        return "Número"

    return "Texto"


def construir_metricas_numericas(df, columnas_numericas):
    filas = []

    for columna in columnas_numericas:
        serie = serie_numerica(df[columna])
        filas.append(
            {
                "Columna": columna,
                "Valores_Validos": int(serie.notna().sum()),
                "Suma": round(float(serie.sum()), 2) if serie.notna().any() else 0,
                "Media": round(float(serie.mean()), 2) if serie.notna().any() else 0,
                "Minimo": round(float(serie.min()), 2) if serie.notna().any() else 0,
                "Maximo": round(float(serie.max()), 2) if serie.notna().any() else 0,
            }
        )

    return pd.DataFrame(filas)


def construir_top_categorias(df, columnas_categoricas_detectadas, columna_importe):
    tablas = {}

    for columna in columnas_categoricas_detectadas[:6]:
        temporal = df[[columna]].copy()
        temporal[columna] = temporal[columna].fillna("SIN DATO").astype(str).str.strip()

        if columna_importe:
            temporal["_importe"] = serie_numerica(df[columna_importe])
            top = (
                temporal.groupby(columna, dropna=False)["_importe"]
                .sum()
                .sort_values(ascending=False)
                .head(MAX_TOP_CATEGORIAS)
                .reset_index()
                .rename(columns={"_importe": f"Suma_{columna_importe}"})
            )
        else:
            top = (
                temporal[columna]
                .value_counts(dropna=False)
                .head(MAX_TOP_CATEGORIAS)
                .reset_index()
            )
            top.columns = [columna, "Registros"]

        tablas[columna] = top

    return tablas


def construir_evolucion_temporal(df, columna_fecha, columna_importe):
    if not columna_fecha:
        return pd.DataFrame()

    temporal = pd.DataFrame()
    temporal["Fecha"] = parsear_fechas(df[columna_fecha])
    temporal = temporal.dropna(subset=["Fecha"])

    if temporal.empty:
        return pd.DataFrame()

    temporal["Mes"] = temporal["Fecha"].dt.to_period("M").astype(str)

    if columna_importe:
        temporal["Importe"] = serie_numerica(df.loc[temporal.index, columna_importe])
        evolucion = temporal.groupby("Mes", as_index=False).agg(
            Registros=("Fecha", "count"),
            Importe_Total=("Importe", "sum"),
        )
        evolucion["Importe_Total"] = evolucion["Importe_Total"].round(2)
    else:
        evolucion = temporal.groupby("Mes", as_index=False).agg(Registros=("Fecha", "count"))

    return evolucion.sort_values("Mes").reset_index(drop=True)


def analizar_dashboard(df):
    df_limpio = df.copy().dropna(how="all").reset_index(drop=True)

    columnas_fecha = detectar_columnas_fecha(df_limpio)
    columnas_numericas = detectar_columnas_numericas(df_limpio)
    columnas_cat = columnas_categoricas(df_limpio, columnas_fecha, columnas_numericas)
    columna_importe = elegir_columna_importe(df_limpio, columnas_numericas)
    columna_fecha = columnas_fecha[0] if columnas_fecha else None

    calidad = construir_calidad(df_limpio)
    metricas = construir_metricas_numericas(df_limpio, columnas_numericas)
    tops = construir_top_categorias(df_limpio, columnas_cat, columna_importe)
    evolucion = construir_evolucion_temporal(df_limpio, columna_fecha, columna_importe)

    resumen = {
        "fecha_analisis": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "registros": len(df_limpio),
        "columnas": len(df_limpio.columns),
        "columnas_fecha": columnas_fecha,
        "columnas_numericas": columnas_numericas,
        "columnas_categoricas": columnas_cat,
        "columna_fecha_principal": columna_fecha,
        "columna_importe_principal": columna_importe,
        "valores_vacios": int(df_limpio.isna().sum().sum()),
        "filas_duplicadas": int(df_limpio.duplicated().sum()),
    }

    return {
        "resumen": resumen,
        "calidad": calidad,
        "metricas": metricas,
        "tops": tops,
        "evolucion": evolucion,
        "base": df_limpio,
        "base_normalizada": construir_base_normalizada(df_limpio, columnas_numericas),
    }


def construir_base_normalizada(df, columnas_numericas):
    base = df.copy()

    for columna in columnas_numericas:
        base[f"{columna}_Normalizado"] = serie_numerica(df[columna])

    return base


def escribir_dataframe(writer, df, nombre):
    df.to_excel(writer, index=False, sheet_name=nombre)
    hoja = writer.sheets[nombre]

    cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    blanco = Font(color="FFFFFF", bold=True)

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    for celda in hoja[1]:
        celda.fill = cabecera
        celda.font = blanco

    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        max_largo = max(len(str(celda.value or "")) for celda in columna)
        hoja.column_dimensions[letra].width = min(max(max_largo + 2, 12), 60)


def escribir_resumen(writer, resultado):
    resumen = resultado["resumen"]
    datos = pd.DataFrame(
        [
            ["Fecha de análisis", resumen["fecha_analisis"]],
            ["Registros analizados", resumen["registros"]],
            ["Columnas detectadas", resumen["columnas"]],
            ["Valores vacíos", resumen["valores_vacios"]],
            ["Filas duplicadas exactas", resumen["filas_duplicadas"]],
            ["Columna fecha principal", resumen["columna_fecha_principal"] or "No detectada"],
            ["Columna importe principal", resumen["columna_importe_principal"] or "No detectada"],
            ["Columnas numéricas", ", ".join(map(str, resumen["columnas_numericas"])) or "No detectadas"],
            ["Columnas categóricas", ", ".join(map(str, resumen["columnas_categoricas"][:8])) or "No detectadas"],
            ["Recomendación", "Revisar calidad de datos antes de usar el dashboard para decisiones finales"],
        ],
        columns=["Métrica", "Valor"],
    )

    datos.to_excel(writer, index=False, sheet_name="Resumen")
    hoja = writer.sheets["Resumen"]
    hoja.insert_rows(1, 2)
    hoja["A1"] = "DataSanity Dashboard Generator - Resumen Ejecutivo"
    hoja["A1"].fill = PatternFill(start_color="102033", end_color="102033", fill_type="solid")
    hoja["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    hoja["A1"].alignment = Alignment(horizontal="left")
    hoja.merge_cells("A1:B1")

    for celda in hoja[3]:
        celda.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        celda.font = Font(color="FFFFFF", bold=True)

    for fila in hoja.iter_rows(min_row=4, max_row=hoja.max_row):
        fila[0].font = Font(color="102033", bold=True)

    hoja.freeze_panes = "A4"
    hoja.column_dimensions["A"].width = 32
    hoja.column_dimensions["B"].width = 95


def escribir_tops(writer, tops):
    if not tops:
        pd.DataFrame([["Sin categorías detectadas"]], columns=["Mensaje"]).to_excel(
            writer, index=False, sheet_name="Top_Categorias"
        )
        return

    sheet_name = "Top_Categorias"
    fila_inicio = 1
    workbook = writer.book
    hoja = workbook.create_sheet(sheet_name)
    writer.sheets[sheet_name] = hoja

    cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    blanco = Font(color="FFFFFF", bold=True)

    for nombre_columna, tabla in tops.items():
        hoja.cell(fila_inicio, 1, f"Top por {nombre_columna}")
        hoja.cell(fila_inicio, 1).font = Font(bold=True, size=12)

        for col_idx, columna in enumerate(tabla.columns, start=1):
            celda = hoja.cell(fila_inicio + 1, col_idx, columna)
            celda.fill = cabecera
            celda.font = blanco

        for row_idx, fila in enumerate(tabla.itertuples(index=False), start=fila_inicio + 2):
            for col_idx, valor in enumerate(fila, start=1):
                hoja.cell(row_idx, col_idx, valor)

        if len(tabla) >= 2:
            chart = BarChart()
            chart.title = f"Top {nombre_columna}"
            chart.height = 7
            chart.width = 14
            data = Reference(hoja, min_col=2, min_row=fila_inicio + 1, max_row=fila_inicio + 1 + len(tabla))
            cats = Reference(hoja, min_col=1, min_row=fila_inicio + 2, max_row=fila_inicio + 1 + len(tabla))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            hoja.add_chart(chart, f"D{fila_inicio + 1}")

        fila_inicio += len(tabla) + 18

    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        max_largo = max(len(str(celda.value or "")) for celda in columna)
        hoja.column_dimensions[letra].width = min(max(max_largo + 2, 12), 55)


def escribir_evolucion(writer, evolucion):
    escribir_dataframe(writer, evolucion, "Evolucion_Temporal")

    if evolucion.empty or len(evolucion) < 2:
        return

    hoja = writer.sheets["Evolucion_Temporal"]
    chart = LineChart()
    chart.title = "Evolución temporal"
    chart.height = 8
    chart.width = 18

    max_col = hoja.max_column
    data = Reference(hoja, min_col=2, max_col=max_col, min_row=1, max_row=hoja.max_row)
    cats = Reference(hoja, min_col=1, min_row=2, max_row=hoja.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    hoja.add_chart(chart, "E2")


def exportar_dashboard_excel(resultado):
    output = pd.io.common.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        escribir_resumen(writer, resultado)
        escribir_dataframe(writer, resultado["calidad"], "Calidad_Datos")
        escribir_dataframe(writer, resultado["metricas"], "Metricas_Numericas")
        escribir_tops(writer, resultado["tops"])
        escribir_evolucion(writer, resultado["evolucion"])
        escribir_dataframe(writer, resultado["base_normalizada"].head(MAX_FILAS_EXPORT_BASE), "Base_Normalizada")
        escribir_dataframe(writer, resultado["base"].head(MAX_FILAS_EXPORT_BASE), "Base_Original")

    output.seek(0)
    return output.getvalue()
