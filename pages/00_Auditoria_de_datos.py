import io

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from limpiador import COLUMNAS_ESPERADAS, procesar_auditoria


st.set_page_config(
    page_title="DataSanity - Auditoría de Datos",
    layout="wide",
)

st.markdown(
    """
    <style>

        .stApp {
            background: #f3f8fd;
        }
        [data-testid="stAppViewContainer"] {
            background: #f3f8fd;
        }
        .main {
            background: #f3f8fd;
        }
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
            max-width: 1280px;
        }
        .ds-hero {
            border: 1px solid #e6eaf0;
            border-radius: 10px;
            padding: 24px 26px;
            background: #ffffff;
            margin-bottom: 20px;
        }
        .ds-kicker {
            color: #4b6b8a;
            font-size: 0.84rem;
            font-weight: 700;
            letter-spacing: 0.04em;
            text-transform: uppercase;
            margin-bottom: 6px;
        }
        .ds-title {
            color: #102033;
            font-size: 2.1rem;
            font-weight: 760;
            line-height: 1.16;
            margin-bottom: 8px;
        }
        .ds-subtitle {
            color: #4c5f73;
            font-size: 1.02rem;
            line-height: 1.55;
            max-width: 900px;
            margin-bottom: 0;
        }
        .ds-panel {
            border: 1px solid #e6eaf0;
            border-radius: 10px;
            padding: 16px 18px;
            background: #fbfcfe;
            height: 100%;
        }
        .ds-panel-title {
            color: #102033;
            font-weight: 720;
            margin-bottom: 6px;
        }
        .ds-panel-copy {
            color: #5a6b7c;
            font-size: 0.92rem;
            line-height: 1.45;
            margin: 0;
        }
        div[data-testid="stMetric"] {
            background: #f3f8fd;
            border: 1px solid #cfe0f1;
            border-radius: 10px;
            padding: 14px 16px;
        }
        div[data-testid="stMetric"] label,
        div[data-testid="stMetric"] [data-testid="stMetricValue"] {
            color: #102033;
        }
        div[data-testid="stMetric"] label {
            opacity: 0.72;
        }
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stButton"] button {
            border-radius: 8px;
            font-weight: 700;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="ds-hero">
        <div class="ds-kicker">DataSanity Data Audit</div>
        <div class="ds-title">Audita, limpia y prepara bases de clientes para revisión profesional</div>
        <p class="ds-subtitle">
            Sube un Excel o CSV, identifica datos incompletos o sospechosos y descarga una versión auditada
            con alertas visuales, resumen de incidencias y formato listo para entregar.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

info_col1, info_col2, info_col3 = st.columns(3)
with info_col1:
    st.markdown(
        """
        <div class="ds-panel">
            <div class="ds-panel-title">Validación automática</div>
            <p class="ds-panel-copy">Revisa campos esperados, valores vacíos y datos que necesitan atención.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
with info_col2:
    st.markdown(
        """
        <div class="ds-panel">
            <div class="ds-panel-title">Alertas claras</div>
            <p class="ds-panel-copy">Marca incidencias con colores y etiquetas para acelerar la revisión.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
with info_col3:
    st.markdown(
        """
        <div class="ds-panel">
            <div class="ds-panel-title">Excel descargable</div>
            <p class="ds-panel-copy">Exporta la base auditada junto con un resumen profesional de alertas.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

st.markdown("---")


def leer_archivo(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        return pd.read_csv(archivo)
    return pd.read_excel(archivo)


def color_celdas(valor):
    if isinstance(valor, str) and valor.startswith("REVISAR:"):
        return "background-color: #ffcccc; color: #cc0000; font-weight: bold;"
    if valor == "SIN NOMBRE":
        return "background-color: #fff2cc; color: #7f6000; font-weight: bold;"
    if valor == "GRATIS":
        return "background-color: #d9ead3; color: #274e13; font-weight: bold;"
    return ""


def crear_excel_descargable(df_procesado, resumen):
    output = io.BytesIO()
    resumen_df = pd.DataFrame(
        {
            "Columna": list(resumen.keys()),
            "Alertas": list(resumen.values()),
        }
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_procesado.to_excel(writer, index=False, sheet_name="Base_Limpia")
        resumen_df.to_excel(writer, index=False, sheet_name="Resumen_Auditoria")

        hoja_datos = writer.sheets["Base_Limpia"]
        hoja_resumen = writer.sheets["Resumen_Auditoria"]

        rojo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        amarillo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        verde = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        blanco_negrita = Font(color="FFFFFF", bold=True)

        for hoja in [hoja_datos, hoja_resumen]:
            hoja.freeze_panes = "A2"
            hoja.auto_filter.ref = hoja.dimensions

            for celda in hoja[1]:
                celda.fill = cabecera
                celda.font = blanco_negrita

            for columna in hoja.columns:
                max_largo = 0
                letra = get_column_letter(columna[0].column)

                for celda in columna:
                    valor = "" if celda.value is None else str(celda.value)
                    max_largo = max(max_largo, len(valor))

                    if isinstance(celda.value, str) and celda.value.startswith("REVISAR:"):
                        celda.fill = rojo
                        celda.font = Font(color="CC0000", bold=True)
                    elif celda.value == "SIN NOMBRE":
                        celda.fill = amarillo
                        celda.font = Font(color="7F6000", bold=True)
                    elif celda.value == "GRATIS":
                        celda.fill = verde
                        celda.font = Font(color="274E13", bold=True)

                hoja.column_dimensions[letra].width = min(max(max_largo + 2, 12), 55)

    output.seek(0)
    return output.getvalue()


st.markdown("### Carga y análisis")
st.caption("Sube tu archivo Excel o CSV para iniciar la auditoría.")
archivo_cargado = st.file_uploader(
    "Archivo Excel o CSV",
    type=["xlsx", "xls", "csv"],
)

if archivo_cargado is not None:
    try:
        df_original = leer_archivo(archivo_cargado)
    except Exception as error:
        st.error(f"No se pudo leer el archivo: {error}")
        st.stop()

    if df_original.empty:
        st.error("El archivo está vacío.")
        st.stop()

    st.subheader("Vista previa de los datos originales")
    st.dataframe(df_original.head(20), width="stretch")

    columnas_faltantes = [
        columna for columna in COLUMNAS_ESPERADAS
        if columna not in df_original.columns
    ]

    if columnas_faltantes:
        st.warning(
            "Faltan columnas esperadas: "
            + ", ".join(columnas_faltantes)
            + ". La app limpiará solo las columnas reconocidas."
        )

    if st.button("Ejecutar auditoría de datos"):
        df_procesado, total, total_alertas, resumen = procesar_auditoria(df_original)
        st.session_state["df_procesado"] = df_procesado
        st.session_state["total"] = total
        st.session_state["total_alertas"] = total_alertas
        st.session_state["resumen"] = resumen
else:
    st.info("Sube un archivo en el bloque de carga para iniciar el análisis.")

if "df_procesado" in st.session_state:
    df_procesado = st.session_state["df_procesado"]
    total = st.session_state["total"]
    total_alertas = st.session_state["total_alertas"]
    resumen = st.session_state["resumen"]

    filas_con_alertas = int(
        df_procesado.astype(str)
        .apply(lambda fila: fila.str.startswith("REVISAR:").any(), axis=1)
        .sum()
    )

    st.success("Auditoría completada.")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Registros analizados", total)
    col2.metric("Filas con alertas", filas_con_alertas)
    col3.metric("Celdas a revisar", total_alertas)
    col4.metric("Columnas auditadas", len(resumen))

    if resumen:
        st.subheader("Resumen de alertas por columna")
        st.dataframe(
            pd.DataFrame(
                {
                    "Columna": list(resumen.keys()),
                    "Alertas": list(resumen.values()),
                }
            ),
            width="stretch",
            hide_index=True,
        )

    st.subheader("Base de datos auditada")

    formatos_columnas = {}
    if "Facturacion_Mensual" in df_procesado.columns:
        formatos_columnas["Facturacion_Mensual"] = (
            lambda valor: f"{valor:.2f}" if isinstance(valor, (int, float)) else str(valor)
        )

    try:
        df_estilizado = df_procesado.style.map(color_celdas).format(formatos_columnas)
    except AttributeError:
        df_estilizado = df_procesado.style.applymap(color_celdas).format(formatos_columnas)

    st.dataframe(df_estilizado, width="stretch")

    archivo_excel = crear_excel_descargable(df_procesado, resumen)

    st.download_button(
        label="Descargar reporte auditado (.xlsx)",
        data=archivo_excel,
        file_name="reporte_auditoria_datos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
