import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from buscador_duplicados import analizar_duplicados


st.set_page_config(
    page_title="DataSanity - Duplicate Finder",
    layout="wide",
)

st.markdown(
    """
    <style>

        .stApp {
            background: #fff7ef;
        }
        [data-testid="stAppViewContainer"] {
            background: #fff7ef;
        }
        .main {
            background: #fff7ef;
        }
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
            max-width: 1280px;
        }
        .ds-hero {
            border: 1px solid #e6eaf0;
            border-radius: 10px;
            padding: 24px 26px;
            background: #ffffff;
            margin-bottom: 20px;
        }
        .ds-kicker {
            color: #4b6b8a;
            font-size: 0.84rem;
            font-weight: 700;
            letter-spacing: 0.04em;
            text-transform: uppercase;
            margin-bottom: 6px;
        }
        .ds-title {
            color: #102033;
            font-size: 2.1rem;
            font-weight: 760;
            line-height: 1.16;
            margin-bottom: 8px;
        }
        .ds-subtitle {
            color: #4c5f73;
            font-size: 1.02rem;
            line-height: 1.55;
            max-width: 880px;
            margin-bottom: 0;
        }
        .ds-panel {
            border: 1px solid #e6eaf0;
            border-radius: 10px;
            padding: 16px 18px;
            background: #fbfcfe;
            height: 100%;
        }
        .ds-panel-title {
            color: #102033;
            font-weight: 720;
            margin-bottom: 6px;
        }
        .ds-panel-copy {
            color: #5a6b7c;
            font-size: 0.92rem;
            line-height: 1.45;
            margin: 0;
        }
        .ds-legend {
            display: grid;
            grid-template-columns: 1fr;
            gap: 8px;
            margin: 10px 0 18px 0;
        }
        .ds-pill {
            display: block;
            border-radius: 8px;
            padding: 8px 10px;
            font-size: 0.82rem;
            font-weight: 720;
            border: 1px solid transparent;
            line-height: 1.25;
        }
        .ds-high {
            background: #d9ead3;
            color: #274e13;
            border-color: #b7d7a8;
        }
        .ds-medium {
            background: #fff2cc;
            color: #7f6000;
            border-color: #f1d98a;
        }
        .ds-low {
            background: #fce5cd;
            color: #783f04;
            border-color: #edc39b;
        }
        div[data-testid="stMetric"] {
            background: #20242c;
            border: 1px solid #353b46;
            border-radius: 10px;
            padding: 14px 16px;
        }
        div[data-testid="stMetric"] label,
        div[data-testid="stMetric"] [data-testid="stMetricValue"] {
            color: #f4f7fb;
        }
        div[data-testid="stMetric"] label {
            opacity: 0.82;
        }
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stButton"] button {
            border-radius: 8px;
            font-weight: 700;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="ds-hero">
        <div class="ds-kicker">DataSanity Duplicate Finder</div>
        <div class="ds-title">Detecta, agrupa y prepara clientes duplicados para revisión</div>
        <p class="ds-subtitle">
            Analiza bases de clientes en Excel o CSV, encuentra coincidencias por email, teléfono y nombres parecidos,
            y genera un reporte profesional con pares detectados, grupos y una propuesta de fusión revisable.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

info_col1, info_col2, info_col3 = st.columns(3)
with info_col1:
    st.markdown(
        """
        <div class="ds-panel">
            <div class="ds-panel-title">Detección</div>
            <p class="ds-panel-copy">Cruza emails, teléfonos normalizados y similitud de nombres para encontrar duplicados probables.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
with info_col2:
    st.markdown(
        """
        <div class="ds-panel">
            <div class="ds-panel-title">Revisión</div>
            <p class="ds-panel-copy">Clasifica cada coincidencia por confianza y muestra el motivo de la detección.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
with info_col3:
    st.markdown(
        """
        <div class="ds-panel">
            <div class="ds-panel-title">Entrega</div>
            <p class="ds-panel-copy">Exporta un Excel con resumen, pares, grupos, base marcada y fusión sugerida.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def leer_archivo(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        return pd.read_csv(archivo)
    return pd.read_excel(archivo)


def color_confianza(valor):
    if valor == "ALTA":
        return "background-color: #d9ead3; color: #274e13; font-weight: bold;"
    if valor == "MEDIA":
        return "background-color: #fff2cc; color: #7f6000; font-weight: bold;"
    if valor == "BAJA":
        return "background-color: #fce5cd; color: #783f04; font-weight: bold;"
    return ""


def color_grupo(valor):
    if pd.notna(valor) and str(valor).strip() not in {"", "nan"}:
        return "background-color: #d9eaf7; font-weight: bold;"
    return ""


def escribir_hoja(writer, df, nombre):
    df.to_excel(writer, index=False, sheet_name=nombre)
    hoja = writer.sheets[nombre]

    cabecera = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    blanco_negrita = Font(color="FFFFFF", bold=True)
    verde = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    amarillo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    azul = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    for celda in hoja[1]:
        celda.fill = cabecera
        celda.font = blanco_negrita

    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        max_largo = 0
        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            max_largo = max(max_largo, len(valor))

            if celda.value == "ALTA":
                celda.fill = verde
                celda.font = Font(color="274E13", bold=True)
            elif celda.value == "MEDIA":
                celda.fill = amarillo
                celda.font = Font(color="7F6000", bold=True)
            elif celda.column_letter and hoja.cell(1, celda.column).value == "Grupo_Duplicado" and valor not in {"", "None", "nan"}:
                celda.fill = azul

        hoja.column_dimensions[letra].width = min(max(max_largo + 2, 12), 60)


def construir_resumen_ejecutivo(resultado):
    resumen = resultado["resumen"]
    pares = resultado["pares"]

    alta = int((pares["Confianza"] == "ALTA").sum()) if not pares.empty else 0
    media = int((pares["Confianza"] == "MEDIA").sum()) if not pares.empty else 0
    baja = int((pares["Confianza"] == "BAJA").sum()) if not pares.empty else 0
    score_medio = round(float(pares["Score"].mean()), 1) if not pares.empty else 0
    score_max = int(pares["Score"].max()) if not pares.empty else 0

    return pd.DataFrame(
        [
            ["Fecha de auditoría", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Registros analizados", resumen["registros_analizados"]],
            ["Duplicados potenciales", resumen["pares_detectados"]],
            ["Grupos de clientes detectados", resumen["grupos_detectados"]],
            ["Registros implicados", resumen["registros_implicados"]],
            ["Pares confianza ALTA", alta],
            ["Pares confianza MODERADA", media],
            ["Pares confianza BAJA", baja],
            ["Score medio de coincidencias", score_medio],
            ["Score máximo detectado", score_max],
            ["Recomendación", "Revisar primero grupos con confianza ALTA y registros con varios datos en conflicto"],
            ["Criterio de seguridad", "El reporte no fusiona datos de forma irreversible; propone una consolidación revisable"],
        ],
        columns=["Métrica", "Valor"],
    )


def construir_metodologia():
    return pd.DataFrame(
        [
            ["Email exacto", "Coincidencia fuerte cuando dos registros comparten el mismo email válido."],
            ["Teléfono exacto", "Coincidencia fuerte tras normalizar prefijos, espacios, puntos y formato español."],
            ["Nombre parecido", "Comparación de similitud para detectar variaciones, tildes, guiones o abreviaturas."],
            ["Score", "Puntuación de 0 a 100 que resume la fuerza de la coincidencia."],
            ["Confianza ALTA", "Probable duplicado. Se recomienda revisar y fusionar si los datos de negocio encajan."],
            ["Confianza MODERADA", "Posible duplicado. Revisar manualmente antes de fusionar."],
            ["Confianza BAJA", "Coincidencia débil. Puede servir para auditoría exploratoria, no para fusión automática."],
            ["Fusión sugerida", "Conserva valores útiles. Si hay conflicto, separa alternativas con | para revisión."],
            ["Limitación", "La herramienta no inventa datos y no elimina registros originales automáticamente."],
        ],
        columns=["Elemento", "Descripción"],
    )


def construir_acciones_recomendadas(resultado):
    resumen = resultado["resumen"]
    pares = resultado["pares"]
    acciones = []

    if resumen["pares_detectados"] == 0:
        acciones.append(["Sin duplicados", "No se detectaron duplicados con la sensibilidad actual."])
    else:
        acciones.append(["Prioridad 1", "Revisar la hoja Grupos_Duplicados y validar los grupos con mayor Score_Maximo."])
        acciones.append(["Prioridad 2", "Abrir Fusion_Sugerida y comprobar campos con valores separados por |."])
        acciones.append(["Prioridad 3", "Usar Base_Marcada para localizar todos los registros originales implicados."])

    if not pares.empty and int((pares["Confianza"] == "ALTA").sum()) > 0:
        acciones.append(["Confianza alta", "Los pares ALTA son los mejores candidatos para consolidación."])

    acciones.append(["Control final", "Conservar una copia de la base original antes de aplicar fusiones definitivas."])

    return pd.DataFrame(acciones, columns=["Acción", "Detalle"])


def aplicar_formato_resumen(writer, nombre):
    hoja = writer.sheets[nombre]
    titulo_fill = PatternFill(start_color="102033", end_color="102033", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white = Font(color="FFFFFF", bold=True)
    label_font = Font(color="102033", bold=True)

    hoja.insert_rows(1, 2)
    hoja["A1"] = "DataSanity Duplicate Finder - Reporte de auditoría"
    hoja["A1"].fill = titulo_fill
    hoja["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    hoja["A1"].alignment = Alignment(horizontal="left")
    hoja.merge_cells("A1:B1")

    for celda in hoja[3]:
        celda.fill = header_fill
        celda.font = white

    for fila in hoja.iter_rows(min_row=4, max_row=hoja.max_row):
        fila[0].font = label_font
        if str(fila[0].value).startswith("Pares confianza"):
            if "ALTA" in str(fila[0].value):
                fila[1].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            elif "MODERADA" in str(fila[0].value):
                fila[1].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif "BAJA" in str(fila[0].value):
                fila[1].fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")

    hoja.freeze_panes = "A4"
    hoja.column_dimensions["A"].width = 34
    hoja.column_dimensions["B"].width = 95


def crear_excel_resultado(resultado):
    output = io.BytesIO()
    resumen_df = construir_resumen_ejecutivo(resultado)
    metodologia_df = construir_metodologia()
    acciones_df = construir_acciones_recomendadas(resultado)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        escribir_hoja(writer, resumen_df, "Resumen")
        aplicar_formato_resumen(writer, "Resumen")
        escribir_hoja(writer, acciones_df, "Acciones_Recomendadas")
        escribir_hoja(writer, metodologia_df, "Metodologia")
        escribir_hoja(writer, resultado["pares"], "Pares_Detectados")
        escribir_hoja(writer, resultado["grupos"], "Grupos_Duplicados")
        escribir_hoja(writer, resultado["fusion_sugerida"], "Fusion_Sugerida")
        escribir_hoja(writer, resultado["base_marcada"], "Base_Marcada")

    output.seek(0)
    return output.getvalue()


st.markdown("### Carga y análisis")
st.caption("Sube tu archivo Excel o CSV y ajusta la sensibilidad antes de buscar duplicados.")

upload_col, config_col = st.columns([1.25, 1])
with upload_col:
    archivo_cargado = st.file_uploader(
        "Archivo Excel o CSV",
        type=["xlsx", "xls", "csv"],
        help="Sube una base con columnas como nombre, email, teléfono o ID de cliente.",
    )
with config_col:
    umbral = st.slider(
        "Sensibilidad de detección",
        min_value=70,
        max_value=95,
        value=78,
        help=(
            "Controla el score mínimo para mostrar una coincidencia. "
            "70 detecta más posibles duplicados; 95 muestra solo coincidencias muy seguras."
        ),
    )
    st.markdown(
        """
        <div class="ds-legend">
            <span class="ds-pill ds-low">70-77 · detección amplia: encuentra más posibles duplicados</span>
            <span class="ds-pill ds-medium">78-89 · equilibrio recomendado: revisión moderada</span>
            <span class="ds-pill ds-high">90-95 · detección estricta: solo coincidencias muy seguras</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

if archivo_cargado is not None:
    try:
        df_original = leer_archivo(archivo_cargado)
    except Exception as error:
        st.error(f"No se pudo leer el archivo: {error}")
        st.stop()

    if df_original.empty:
        st.error("El archivo está vacío.")
        st.stop()

    st.subheader("Vista previa de la base original")
    st.dataframe(df_original.head(20), width="stretch")

    if st.button("Buscar duplicados", type="primary"):
        resultado = analizar_duplicados(df_original, umbral=umbral)
        st.session_state["resultado_duplicados"] = resultado
else:
    st.info("Sube un archivo en el bloque de carga para iniciar el análisis.")

if "resultado_duplicados" in st.session_state:
    resultado = st.session_state["resultado_duplicados"]
    resumen = resultado["resumen"]

    st.success("Análisis de duplicados completado.")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Registros analizados", resumen["registros_analizados"])
    col2.metric("Duplicados potenciales", resumen["pares_detectados"])
    col3.metric("Grupos detectados", resumen["grupos_detectados"])
    col4.metric("Registros a revisar", resumen["registros_implicados"])

    st.markdown(
        """
        <div class="ds-legend">
            <span class="ds-pill ds-high">ALTA = probable duplicado</span>
            <span class="ds-pill ds-medium">MODERADA = revisar antes de fusionar</span>
            <span class="ds-pill ds-low">BAJA = coincidencia débil</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    tab_resumen, tab_pares, tab_grupos, tab_fusion, tab_base = st.tabs(
        [
            "Resumen",
            "Pares detectados",
            "Grupos",
            "Propuesta de fusión",
            "Base marcada",
        ]
    )

    with tab_resumen:
        st.subheader("Resumen ejecutivo")
        resumen_tabla = pd.DataFrame(
            [
                ["Registros analizados", resumen["registros_analizados"]],
                ["Duplicados potenciales", resumen["pares_detectados"]],
                ["Grupos detectados", resumen["grupos_detectados"]],
                ["Registros a revisar", resumen["registros_implicados"]],
            ],
            columns=["Métrica", "Valor"],
        )
        st.dataframe(resumen_tabla, width="stretch", hide_index=True)
        st.write(
            "El reporte no fusiona datos de forma irreversible. Agrupa coincidencias y propone una consolidación "
            "para que puedas revisarla antes de tomar decisiones."
        )

    with tab_pares:
        st.subheader("Pares detectados")
        if resultado["pares"].empty:
            st.info("No se han detectado duplicados con el umbral actual.")
        else:
            try:
                pares_estilizados = resultado["pares"].style.map(color_confianza, subset=["Confianza"])
            except AttributeError:
                pares_estilizados = resultado["pares"].style.applymap(color_confianza, subset=["Confianza"])
            st.dataframe(pares_estilizados, width="stretch", hide_index=True)

    with tab_grupos:
        st.subheader("Grupos de clientes duplicados")
        if resultado["grupos"].empty:
            st.info("No hay grupos de duplicados para fusionar.")
        else:
            st.dataframe(resultado["grupos"], width="stretch", hide_index=True)

    with tab_fusion:
        st.subheader("Propuesta de fusión revisable")
        st.write(
            "La propuesta conserva los valores útiles. Cuando hay conflictos, "
            "los deja separados con `|` para revisión humana."
        )
        if resultado["fusion_sugerida"].empty:
            st.info("No hay registros fusionados sugeridos.")
        else:
            st.dataframe(resultado["fusion_sugerida"], width="stretch", hide_index=True)

    with tab_base:
        st.subheader("Base marcada")
        if resultado["base_marcada"].empty:
            st.info("No hay registros para mostrar.")
        else:
            try:
                base_estilizada = resultado["base_marcada"].style.map(color_grupo, subset=["Grupo_Duplicado"])
            except AttributeError:
                base_estilizada = resultado["base_marcada"].style.applymap(color_grupo, subset=["Grupo_Duplicado"])
            st.dataframe(base_estilizada, width="stretch", hide_index=True)

    archivo_excel = crear_excel_resultado(resultado)
    st.download_button(
        label="Descargar reporte profesional de duplicados (.xlsx)",
        data=archivo_excel,
        file_name="reporte_duplicados_clientes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
